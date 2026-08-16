"""Big 4 Tier Multi-Tab Professional Excel Exporter for ICAI Non-Corporate Financial Statements."""
import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from typing import Dict, Any

# Professional Corporate Typography & Colors (Big 4 Navy & Slate Theme)
FONT_COVER = Font(name="Segoe UI", size=16, bold=True, color="1E3A8A")
FONT_TITLE = Font(name="Segoe UI", size=13, bold=True, color="0F172A")
FONT_SUBTITLE = Font(name="Segoe UI", size=10, italic=True, color="475569")
FONT_SECTION_HDR = Font(name="Segoe UI", size=10, bold=True, color="FFFFFF")
FONT_SUB_HDR = Font(name="Segoe UI", size=10, bold=True, color="1E293B")
FONT_ITEM = Font(name="Segoe UI", size=10, color="1E293B")
FONT_TOTAL = Font(name="Segoe UI", size=10, bold=True, color="000000")

FILL_NAVY_DARK = PatternFill(start_color="0F172A", end_color="0F172A", fill_type="solid")
FILL_PRIMARY = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")
FILL_ACCENT_BLUE = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
FILL_SUBTOTAL = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
FILL_HIGHLIGHT = PatternFill(start_color="EFF6FF", end_color="EFF6FF", fill_type="solid")
FILL_GREEN = PatternFill(start_color="ECFDF5", end_color="ECFDF5", fill_type="solid")

BORDER_THIN = Border(
    top=Side(style="thin", color="CBD5E1"),
    bottom=Side(style="thin", color="CBD5E1")
)
BORDER_TOTAL = Border(
    top=Side(style="thin", color="000000"),
    bottom=Side(style="double", color="000000")
)
BORDER_CARD = Border(
    top=Side(style="thin", color="E2E8F0"),
    bottom=Side(style="thin", color="E2E8F0"),
    left=Side(style="thin", color="E2E8F0"),
    right=Side(style="thin", color="E2E8F0")
)

ALIGN_LEFT = Alignment(horizontal="left", vertical="center")
ALIGN_CENTER = Alignment(horizontal="center", vertical="center")
ALIGN_RIGHT = Alignment(horizontal="right", vertical="center")

NUM_FORMAT = "₹ #,##0.00;[Red](₹ #,##0.00);\"-\""

def apply_row_style(ws, row_idx, fill=None, font=None, border=None, num_format=None, align_cols=None):
    """Applies styles to all cells in a row."""
    for col_idx in range(1, ws.max_column + 1):
        cell = ws.cell(row=row_idx, column=col_idx)
        if fill:
            cell.fill = fill
        if font:
            cell.font = font
        if border:
            cell.border = border
        if num_format and col_idx in [3, 4, 5, 6, 7, 8]:
            cell.number_format = num_format
        if align_cols and col_idx in align_cols:
            cell.alignment = align_cols[col_idx]

def create_excel_workbook(statement_data: Dict[str, Any]) -> io.BytesIO:
    """Generates a complete 10-sheet Big 4 audit grade Excel workbook."""
    wb = openpyxl.Workbook()
    wb.remove(wb.active) # Remove initial blank sheet

    entity = statement_data.get("entity", {})
    entity_lvl = statement_data.get("entity_level", {})
    bs_data = statement_data.get("balance_sheet", {})
    pnl_data = statement_data.get("profit_and_loss", {})
    cfs_data = statement_data.get("cash_flow_statement", {})
    notes_data = statement_data.get("notes", [])
    ratios = statement_data.get("financial_ratios", [])
    ppe = statement_data.get("ppe_schedule", {})
    aging = statement_data.get("aging_schedules", {})
    policies = statement_data.get("accounting_policies", [])
    audit_rep = statement_data.get("audit_report", {})

    # =========================================================================
    # SHEET 1: EXECUTIVE SUMMARY & AUDIT DASHBOARD
    # =========================================================================
    ws_exec = wb.create_sheet(title="Executive Summary")
    ws_exec.views.sheetView[0].showGridLines = True

    ws_exec.append([entity.get("name", "Non-Corporate Entity")])
    ws_exec.cell(row=1, column=1).font = FONT_COVER
    ws_exec.append([f"EXECUTIVE AUDIT SUMMARY & FINANCIAL HIGHLIGHTS — FY {entity.get('financial_year', '2025-26')}"])
    ws_exec.cell(row=2, column=1).font = FONT_SUBTITLE
    ws_exec.append([])

    ws_exec.append(["Entity Classification & Audit Opinion", "", "Key Financial Metrics", ""])
    apply_row_style(ws_exec, ws_exec.max_row, fill=FILL_PRIMARY, font=FONT_SECTION_HDR)

    eq_tot = bs_data.get("equity_and_liabilities", {}).get("owners_funds", {}).get("total", 0.0)
    ast_tot = bs_data.get("assets", {}).get("total", 0.0)
    rev_tot = pnl_data.get("income", {}).get("total", 0.0)
    pat_tot = pnl_data.get("profit_after_tax", 0.0)

    ws_exec.append(["Entity Name:", entity.get("name"), "Total Income (Revenue):", rev_tot])
    apply_row_style(ws_exec, ws_exec.max_row, font=FONT_ITEM, num_format=NUM_FORMAT)
    ws_exec.append(["Entity Type:", entity.get("entity_type"), "Net Profit After Tax (PAT):", pat_tot])
    apply_row_style(ws_exec, ws_exec.max_row, font=FONT_ITEM, num_format=NUM_FORMAT)
    ws_exec.append(["ICAI Category:", f"{entity_lvl.get('level')} ({entity_lvl.get('category')})", "Owners' Funds (Net Worth):", eq_tot])
    apply_row_style(ws_exec, ws_exec.max_row, font=FONT_ITEM, num_format=NUM_FORMAT)
    ws_exec.append(["Audit Opinion:", audit_rep.get("opinion_type"), "Total Assets / BS Size:", ast_tot])
    apply_row_style(ws_exec, ws_exec.max_row, font=FONT_ITEM, num_format=NUM_FORMAT)
    ws_exec.append([])

    # Ratios summary section in Executive Sheet
    ws_exec.append(["Key Ratio", "Formula / Concept", "Value", "Benchmark", "Health Status"])
    apply_row_style(ws_exec, ws_exec.max_row, fill=FILL_NAVY_DARK, font=FONT_SECTION_HDR)

    for r in ratios[:6]:
        ws_exec.append([r["name"], r["formula"], f"{r['value']} {r['unit']}", r["benchmark"], r["status"]])
        apply_row_style(ws_exec, ws_exec.max_row, font=FONT_ITEM, align_cols={3: ALIGN_CENTER, 4: ALIGN_CENTER, 5: ALIGN_CENTER})

    ws_exec.column_dimensions["A"].width = 32
    ws_exec.column_dimensions["B"].width = 38
    ws_exec.column_dimensions["C"].width = 24
    ws_exec.column_dimensions["D"].width = 24
    ws_exec.column_dimensions["E"].width = 20

    # =========================================================================
    # SHEET 2: BALANCE SHEET
    # =========================================================================
    ws_bs = wb.create_sheet(title="Balance Sheet")
    ws_bs.views.sheetView[0].showGridLines = True

    ws_bs.append([entity.get("name")])
    ws_bs.append([f"BALANCE SHEET AS AT {entity.get('as_on_date', '31-03-2026').upper()}"])
    ws_bs.append([f"(Prepared as per ICAI Technical Guide for {entity.get('entity_type')})"])
    ws_bs.append([])

    ws_bs.append(["Particulars", "", "Note No.", f"As at {entity.get('as_on_date')}"])
    apply_row_style(ws_bs, ws_bs.max_row, fill=FILL_PRIMARY, font=FONT_SECTION_HDR, align_cols={1: ALIGN_LEFT, 3: ALIGN_CENTER, 4: ALIGN_RIGHT})

    eq = bs_data.get("equity_and_liabilities", {})
    assets = bs_data.get("assets", {})

    # I. EQUITY AND LIABILITIES
    ws_bs.append(["I. EQUITY AND LIABILITIES", "", "", ""])
    apply_row_style(ws_bs, ws_bs.max_row, fill=FILL_SUBTOTAL, font=FONT_SUB_HDR)

    # 1. Owners' Funds
    ws_bs.append(["  1. Owners' Funds", "", "", ""])
    apply_row_style(ws_bs, ws_bs.max_row, font=FONT_SUB_HDR)
    of = eq.get("owners_funds", {})
    ws_bs.append(["     (a) Owner's / Partners' Capital Accounts", "", of.get("owners_capital", {}).get("note", "-"), of.get("owners_capital", {}).get("amount", 0.0)])
    apply_row_style(ws_bs, ws_bs.max_row, font=FONT_ITEM, num_format=NUM_FORMAT, align_cols={3: ALIGN_CENTER, 4: ALIGN_RIGHT})
    ws_bs.append(["     (b) Reserves and Surplus", "", of.get("reserves_and_surplus", {}).get("note", "-"), of.get("reserves_and_surplus", {}).get("amount", 0.0)])
    apply_row_style(ws_bs, ws_bs.max_row, font=FONT_ITEM, num_format=NUM_FORMAT, align_cols={3: ALIGN_CENTER, 4: ALIGN_RIGHT})
    ws_bs.append(["     Sub-total: Owners' Funds", "", "", of.get("total", 0.0)])
    apply_row_style(ws_bs, ws_bs.max_row, font=FONT_SUB_HDR, border=BORDER_THIN, num_format=NUM_FORMAT, align_cols={4: ALIGN_RIGHT})

    # 2. Non-Current Liabilities
    ws_bs.append(["  2. Non-Current Liabilities", "", "", ""])
    apply_row_style(ws_bs, ws_bs.max_row, font=FONT_SUB_HDR)
    ncl = eq.get("non_current_liabilities", {})
    ws_bs.append(["     (a) Long-Term Borrowings", "", ncl.get("long_term_borrowings", {}).get("note", "-"), ncl.get("long_term_borrowings", {}).get("amount", 0.0)])
    apply_row_style(ws_bs, ws_bs.max_row, font=FONT_ITEM, num_format=NUM_FORMAT, align_cols={3: ALIGN_CENTER, 4: ALIGN_RIGHT})
    ws_bs.append(["     (b) Deferred Tax Liabilities (Net)", "", ncl.get("deferred_tax_liabilities", {}).get("note", "-"), ncl.get("deferred_tax_liabilities", {}).get("amount", 0.0)])
    apply_row_style(ws_bs, ws_bs.max_row, font=FONT_ITEM, num_format=NUM_FORMAT, align_cols={3: ALIGN_CENTER, 4: ALIGN_RIGHT})
    ws_bs.append(["     (c) Other Long-Term Liabilities", "", ncl.get("other_long_term_liabilities", {}).get("note", "-"), ncl.get("other_long_term_liabilities", {}).get("amount", 0.0)])
    apply_row_style(ws_bs, ws_bs.max_row, font=FONT_ITEM, num_format=NUM_FORMAT, align_cols={3: ALIGN_CENTER, 4: ALIGN_RIGHT})
    ws_bs.append(["     (d) Long-Term Provisions", "", ncl.get("long_term_provisions", {}).get("note", "-"), ncl.get("long_term_provisions", {}).get("amount", 0.0)])
    apply_row_style(ws_bs, ws_bs.max_row, font=FONT_ITEM, num_format=NUM_FORMAT, align_cols={3: ALIGN_CENTER, 4: ALIGN_RIGHT})
    ws_bs.append(["     Sub-total: Non-Current Liabilities", "", "", ncl.get("total", 0.0)])
    apply_row_style(ws_bs, ws_bs.max_row, font=FONT_SUB_HDR, border=BORDER_THIN, num_format=NUM_FORMAT, align_cols={4: ALIGN_RIGHT})

    # 3. Current Liabilities
    ws_bs.append(["  3. Current Liabilities", "", "", ""])
    apply_row_style(ws_bs, ws_bs.max_row, font=FONT_SUB_HDR)
    cl = eq.get("current_liabilities", {})
    ws_bs.append(["     (a) Short-Term Borrowings", "", cl.get("short_term_borrowings", {}).get("note", "-"), cl.get("short_term_borrowings", {}).get("amount", 0.0)])
    apply_row_style(ws_bs, ws_bs.max_row, font=FONT_ITEM, num_format=NUM_FORMAT, align_cols={3: ALIGN_CENTER, 4: ALIGN_RIGHT})
    ws_bs.append(["     (b) Trade Payables", "", cl.get("trade_payables", {}).get("note", "-"), cl.get("trade_payables", {}).get("amount", 0.0)])
    apply_row_style(ws_bs, ws_bs.max_row, font=FONT_ITEM, num_format=NUM_FORMAT, align_cols={3: ALIGN_CENTER, 4: ALIGN_RIGHT})
    ws_bs.append(["     (c) Other Current Liabilities", "", cl.get("other_current_liabilities", {}).get("note", "-"), cl.get("other_current_liabilities", {}).get("amount", 0.0)])
    apply_row_style(ws_bs, ws_bs.max_row, font=FONT_ITEM, num_format=NUM_FORMAT, align_cols={3: ALIGN_CENTER, 4: ALIGN_RIGHT})
    ws_bs.append(["     (d) Short-Term Provisions", "", cl.get("short_term_provisions", {}).get("note", "-"), cl.get("short_term_provisions", {}).get("amount", 0.0)])
    apply_row_style(ws_bs, ws_bs.max_row, font=FONT_ITEM, num_format=NUM_FORMAT, align_cols={3: ALIGN_CENTER, 4: ALIGN_RIGHT})
    ws_bs.append(["     Sub-total: Current Liabilities", "", "", cl.get("total", 0.0)])
    apply_row_style(ws_bs, ws_bs.max_row, font=FONT_SUB_HDR, border=BORDER_THIN, num_format=NUM_FORMAT, align_cols={4: ALIGN_RIGHT})

    # TOTAL EQUITY AND LIABILITIES
    ws_bs.append(["TOTAL - EQUITY AND LIABILITIES", "", "", eq.get("total", 0.0)])
    apply_row_style(ws_bs, ws_bs.max_row, fill=FILL_HIGHLIGHT, font=FONT_TOTAL, border=BORDER_TOTAL, num_format=NUM_FORMAT, align_cols={4: ALIGN_RIGHT})
    ws_bs.append([])

    # II. ASSETS
    ws_bs.append(["II. ASSETS", "", "", ""])
    apply_row_style(ws_bs, ws_bs.max_row, fill=FILL_SUBTOTAL, font=FONT_SUB_HDR)

    # 1. Non-Current Assets
    ws_bs.append(["  1. Non-Current Assets", "", "", ""])
    apply_row_style(ws_bs, ws_bs.max_row, font=FONT_SUB_HDR)
    nca = assets.get("non_current_assets", {})
    ppe_net = nca.get("ppe_net", {})
    ws_bs.append(["     (a) Property, Plant and Equipment (Net)", "", ppe_net.get("note", "-"), ppe_net.get("amount", 0.0)])
    apply_row_style(ws_bs, ws_bs.max_row, font=FONT_ITEM, num_format=NUM_FORMAT, align_cols={3: ALIGN_CENTER, 4: ALIGN_RIGHT})
    ws_bs.append(["     (b) Intangible Assets", "", nca.get("intangible_assets", {}).get("note", "-"), nca.get("intangible_assets", {}).get("amount", 0.0)])
    apply_row_style(ws_bs, ws_bs.max_row, font=FONT_ITEM, num_format=NUM_FORMAT, align_cols={3: ALIGN_CENTER, 4: ALIGN_RIGHT})
    ws_bs.append(["     (c) Non-Current Investments", "", nca.get("non_current_investments", {}).get("note", "-"), nca.get("non_current_investments", {}).get("amount", 0.0)])
    apply_row_style(ws_bs, ws_bs.max_row, font=FONT_ITEM, num_format=NUM_FORMAT, align_cols={3: ALIGN_CENTER, 4: ALIGN_RIGHT})
    ws_bs.append(["     (d) Deferred Tax Assets (Net)", "", nca.get("deferred_tax_assets", {}).get("note", "-"), nca.get("deferred_tax_assets", {}).get("amount", 0.0)])
    apply_row_style(ws_bs, ws_bs.max_row, font=FONT_ITEM, num_format=NUM_FORMAT, align_cols={3: ALIGN_CENTER, 4: ALIGN_RIGHT})
    ws_bs.append(["     (e) Long-Term Loans and Advances", "", nca.get("long_term_loans_advances", {}).get("note", "-"), nca.get("long_term_loans_advances", {}).get("amount", 0.0)])
    apply_row_style(ws_bs, ws_bs.max_row, font=FONT_ITEM, num_format=NUM_FORMAT, align_cols={3: ALIGN_CENTER, 4: ALIGN_RIGHT})
    ws_bs.append(["     (f) Other Non-Current Assets", "", nca.get("other_non_current_assets", {}).get("note", "-"), nca.get("other_non_current_assets", {}).get("amount", 0.0)])
    apply_row_style(ws_bs, ws_bs.max_row, font=FONT_ITEM, num_format=NUM_FORMAT, align_cols={3: ALIGN_CENTER, 4: ALIGN_RIGHT})
    ws_bs.append(["     Sub-total: Non-Current Assets", "", "", nca.get("total", 0.0)])
    apply_row_style(ws_bs, ws_bs.max_row, font=FONT_SUB_HDR, border=BORDER_THIN, num_format=NUM_FORMAT, align_cols={4: ALIGN_RIGHT})

    # 2. Current Assets
    ws_bs.append(["  2. Current Assets", "", "", ""])
    apply_row_style(ws_bs, ws_bs.max_row, font=FONT_SUB_HDR)
    ca = assets.get("current_assets", {})
    ws_bs.append(["     (a) Current Investments", "", ca.get("current_investments", {}).get("note", "-"), ca.get("current_investments", {}).get("amount", 0.0)])
    apply_row_style(ws_bs, ws_bs.max_row, font=FONT_ITEM, num_format=NUM_FORMAT, align_cols={3: ALIGN_CENTER, 4: ALIGN_RIGHT})
    ws_bs.append(["     (b) Inventories", "", ca.get("inventories", {}).get("note", "-"), ca.get("inventories", {}).get("amount", 0.0)])
    apply_row_style(ws_bs, ws_bs.max_row, font=FONT_ITEM, num_format=NUM_FORMAT, align_cols={3: ALIGN_CENTER, 4: ALIGN_RIGHT})
    ws_bs.append(["     (c) Trade Receivables", "", ca.get("trade_receivables", {}).get("note", "-"), ca.get("trade_receivables", {}).get("amount", 0.0)])
    apply_row_style(ws_bs, ws_bs.max_row, font=FONT_ITEM, num_format=NUM_FORMAT, align_cols={3: ALIGN_CENTER, 4: ALIGN_RIGHT})
    ws_bs.append(["     (d) Cash and Bank Balances", "", ca.get("cash_and_bank_balances", {}).get("note", "-"), ca.get("cash_and_bank_balances", {}).get("amount", 0.0)])
    apply_row_style(ws_bs, ws_bs.max_row, font=FONT_ITEM, num_format=NUM_FORMAT, align_cols={3: ALIGN_CENTER, 4: ALIGN_RIGHT})
    ws_bs.append(["     (e) Short-Term Loans and Advances", "", ca.get("short_term_loans_advances", {}).get("note", "-"), ca.get("short_term_loans_advances", {}).get("amount", 0.0)])
    apply_row_style(ws_bs, ws_bs.max_row, font=FONT_ITEM, num_format=NUM_FORMAT, align_cols={3: ALIGN_CENTER, 4: ALIGN_RIGHT})
    ws_bs.append(["     (f) Other Current Assets", "", ca.get("other_current_assets", {}).get("note", "-"), ca.get("other_current_assets", {}).get("amount", 0.0)])
    apply_row_style(ws_bs, ws_bs.max_row, font=FONT_ITEM, num_format=NUM_FORMAT, align_cols={3: ALIGN_CENTER, 4: ALIGN_RIGHT})
    ws_bs.append(["     Sub-total: Current Assets", "", "", ca.get("total", 0.0)])
    apply_row_style(ws_bs, ws_bs.max_row, font=FONT_SUB_HDR, border=BORDER_THIN, num_format=NUM_FORMAT, align_cols={4: ALIGN_RIGHT})

    # TOTAL ASSETS
    ws_bs.append(["TOTAL - ASSETS", "", "", assets.get("total", 0.0)])
    apply_row_style(ws_bs, ws_bs.max_row, fill=FILL_HIGHLIGHT, font=FONT_TOTAL, border=BORDER_TOTAL, num_format=NUM_FORMAT, align_cols={4: ALIGN_RIGHT})

    ws_bs.column_dimensions["A"].width = 46
    ws_bs.column_dimensions["B"].width = 5
    ws_bs.column_dimensions["C"].width = 12
    ws_bs.column_dimensions["D"].width = 24

    # =========================================================================
    # SHEET 3: STATEMENT OF PROFIT AND LOSS
    # =========================================================================
    ws_pnl = wb.create_sheet(title="Profit & Loss")
    ws_pnl.views.sheetView[0].showGridLines = True

    ws_pnl.append([entity.get("name")])
    ws_pnl.append([f"STATEMENT OF PROFIT AND LOSS FOR THE PERIOD ENDED {entity.get('as_on_date').upper()}"])
    ws_pnl.append([f"(Nature of Expense Method — ICAI Non-Corporate Guidelines)"])
    ws_pnl.append([])

    ws_pnl.append(["Particulars", "", "Note No.", f"For the year ended {entity.get('as_on_date')}"])
    apply_row_style(ws_pnl, ws_pnl.max_row, fill=FILL_PRIMARY, font=FONT_SECTION_HDR, align_cols={1: ALIGN_LEFT, 3: ALIGN_CENTER, 4: ALIGN_RIGHT})

    inc = pnl_data.get("income", {})
    exp = pnl_data.get("expenses", {})

    ws_pnl.append(["I. Revenue from Operations", "", inc.get("revenue_from_operations", {}).get("note", "-"), inc.get("revenue_from_operations", {}).get("amount", 0.0)])
    apply_row_style(ws_pnl, ws_pnl.max_row, font=FONT_ITEM, num_format=NUM_FORMAT, align_cols={3: ALIGN_CENTER, 4: ALIGN_RIGHT})
    ws_pnl.append(["II. Other Income", "", inc.get("other_income", {}).get("note", "-"), inc.get("other_income", {}).get("amount", 0.0)])
    apply_row_style(ws_pnl, ws_pnl.max_row, font=FONT_ITEM, num_format=NUM_FORMAT, align_cols={3: ALIGN_CENTER, 4: ALIGN_RIGHT})
    ws_pnl.append(["III. Total Income (I + II)", "", "", inc.get("total", 0.0)])
    apply_row_style(ws_pnl, ws_pnl.max_row, fill=FILL_SUBTOTAL, font=FONT_SUB_HDR, border=BORDER_THIN, num_format=NUM_FORMAT, align_cols={4: ALIGN_RIGHT})
    ws_pnl.append([])

    ws_pnl.append(["IV. Expenses:", "", "", ""])
    apply_row_style(ws_pnl, ws_pnl.max_row, font=FONT_SUB_HDR)
    ws_pnl.append(["  (a) Cost of Materials Consumed", "", exp.get("cost_of_materials", {}).get("note", "-"), exp.get("cost_of_materials", {}).get("amount", 0.0)])
    apply_row_style(ws_pnl, ws_pnl.max_row, font=FONT_ITEM, num_format=NUM_FORMAT, align_cols={3: ALIGN_CENTER, 4: ALIGN_RIGHT})
    ws_pnl.append(["  (b) Purchases of Stock-in-Trade / COGS", "", exp.get("purchases_stock_in_trade", {}).get("note", "-"), exp.get("purchases_stock_in_trade", {}).get("amount", 0.0)])
    apply_row_style(ws_pnl, ws_pnl.max_row, font=FONT_ITEM, num_format=NUM_FORMAT, align_cols={3: ALIGN_CENTER, 4: ALIGN_RIGHT})
    ws_pnl.append(["  (c) Changes in Inventories", "", exp.get("changes_in_inventories", {}).get("note", "-"), exp.get("changes_in_inventories", {}).get("amount", 0.0)])
    apply_row_style(ws_pnl, ws_pnl.max_row, font=FONT_ITEM, num_format=NUM_FORMAT, align_cols={3: ALIGN_CENTER, 4: ALIGN_RIGHT})
    ws_pnl.append(["  (d) Employee Benefits Expense", "", exp.get("employee_benefits", {}).get("note", "-"), exp.get("employee_benefits", {}).get("amount", 0.0)])
    apply_row_style(ws_pnl, ws_pnl.max_row, font=FONT_ITEM, num_format=NUM_FORMAT, align_cols={3: ALIGN_CENTER, 4: ALIGN_RIGHT})
    ws_pnl.append(["  (e) Finance Costs", "", exp.get("finance_costs", {}).get("note", "-"), exp.get("finance_costs", {}).get("amount", 0.0)])
    apply_row_style(ws_pnl, ws_pnl.max_row, font=FONT_ITEM, num_format=NUM_FORMAT, align_cols={3: ALIGN_CENTER, 4: ALIGN_RIGHT})
    ws_pnl.append(["  (f) Depreciation & Amortisation Expense", "", exp.get("depreciation_amortisation", {}).get("note", "-"), exp.get("depreciation_amortisation", {}).get("amount", 0.0)])
    apply_row_style(ws_pnl, ws_pnl.max_row, font=FONT_ITEM, num_format=NUM_FORMAT, align_cols={3: ALIGN_CENTER, 4: ALIGN_RIGHT})
    ws_pnl.append(["  (g) Other Expenses", "", exp.get("other_expenses", {}).get("note", "-"), exp.get("other_expenses", {}).get("amount", 0.0)])
    apply_row_style(ws_pnl, ws_pnl.max_row, font=FONT_ITEM, num_format=NUM_FORMAT, align_cols={3: ALIGN_CENTER, 4: ALIGN_RIGHT})

    ws_pnl.append(["Total Expenses (IV)", "", "", exp.get("total", 0.0)])
    apply_row_style(ws_pnl, ws_pnl.max_row, fill=FILL_SUBTOTAL, font=FONT_SUB_HDR, border=BORDER_THIN, num_format=NUM_FORMAT, align_cols={4: ALIGN_RIGHT})
    ws_pnl.append([])

    ws_pnl.append(["V. Profit / (Loss) Before Tax (III - IV)", "", "", pnl_data.get("profit_before_tax", 0.0)])
    apply_row_style(ws_pnl, ws_pnl.max_row, font=FONT_SUB_HDR, num_format=NUM_FORMAT, align_cols={4: ALIGN_RIGHT})

    tax = pnl_data.get("tax_expense", {})
    ws_pnl.append(["VI. Tax Expense", "", tax.get("note", "-"), tax.get("amount", 0.0)])
    apply_row_style(ws_pnl, ws_pnl.max_row, font=FONT_ITEM, num_format=NUM_FORMAT, align_cols={3: ALIGN_CENTER, 4: ALIGN_RIGHT})

    ws_pnl.append(["VII. Profit / (Loss) for the period (V - VI)", "", "", pnl_data.get("profit_after_tax", 0.0)])
    apply_row_style(ws_pnl, ws_pnl.max_row, fill=FILL_HIGHLIGHT, font=FONT_TOTAL, border=BORDER_TOTAL, num_format=NUM_FORMAT, align_cols={4: ALIGN_RIGHT})

    ws_pnl.column_dimensions["A"].width = 46
    ws_pnl.column_dimensions["B"].width = 5
    ws_pnl.column_dimensions["C"].width = 12
    ws_pnl.column_dimensions["D"].width = 24

    # =========================================================================
    # SHEET 4: CASH FLOW STATEMENT (AS-3 Indirect)
    # =========================================================================
    ws_cfs = wb.create_sheet(title="Cash Flow Statement")
    ws_cfs.views.sheetView[0].showGridLines = True

    ws_cfs.append([entity.get("name")])
    ws_cfs.append([f"CASH FLOW STATEMENT FOR THE PERIOD ENDED {entity.get('as_on_date').upper()}"])
    ws_cfs.append(["(Prepared as per AS-3 Indirect Method)"])
    ws_cfs.append([])

    ws_cfs.append(["Particulars", "", "", f"Amount ({entity.get('currency_symbol')})"])
    apply_row_style(ws_cfs, ws_cfs.max_row, fill=FILL_PRIMARY, font=FONT_SECTION_HDR, align_cols={1: ALIGN_LEFT, 4: ALIGN_RIGHT})

    cfo = cfs_data.get("operating_activities", {})
    cfi = cfs_data.get("investing_activities", {})
    cff = cfs_data.get("financing_activities", {})
    cfs_sum = cfs_data.get("summary", {})

    ws_cfs.append(["A. CASH FLOW FROM OPERATING ACTIVITIES", "", "", ""])
    apply_row_style(ws_cfs, ws_cfs.max_row, fill=FILL_SUBTOTAL, font=FONT_SUB_HDR)
    ws_cfs.append(["  Net Profit Before Tax", "", "", cfo.get("net_profit_before_tax", 0.0)])
    apply_row_style(ws_cfs, ws_cfs.max_row, font=FONT_ITEM, num_format=NUM_FORMAT, align_cols={4: ALIGN_RIGHT})

    ws_cfs.append(["  Adjustments for Non-Cash & Non-Operating Items:", "", "", ""])
    apply_row_style(ws_cfs, ws_cfs.max_row, font=FONT_SUB_HDR)
    for adj in cfo.get("adjustments", []):
        ws_cfs.append([f"    {adj['label']}", "", "", adj["amount"]])
        apply_row_style(ws_cfs, ws_cfs.max_row, font=FONT_ITEM, num_format=NUM_FORMAT, align_cols={4: ALIGN_RIGHT})

    ws_cfs.append(["  Operating Profit Before Working Capital Changes", "", "", cfo.get("operating_profit_before_wc", 0.0)])
    apply_row_style(ws_cfs, ws_cfs.max_row, font=FONT_SUB_HDR, border=BORDER_THIN, num_format=NUM_FORMAT, align_cols={4: ALIGN_RIGHT})

    ws_cfs.append(["  Working Capital Adjustments:", "", "", ""])
    apply_row_style(ws_cfs, ws_cfs.max_row, font=FONT_SUB_HDR)
    for wc in cfo.get("working_capital_changes", []):
        ws_cfs.append([f"    {wc['label']}", "", "", wc["amount"]])
        apply_row_style(ws_cfs, ws_cfs.max_row, font=FONT_ITEM, num_format=NUM_FORMAT, align_cols={4: ALIGN_RIGHT})

    ws_cfs.append(["  Cash Generated from Operations", "", "", cfo.get("cash_generated_from_operations", 0.0)])
    apply_row_style(ws_cfs, ws_cfs.max_row, font=FONT_SUB_HDR, num_format=NUM_FORMAT, align_cols={4: ALIGN_RIGHT})
    ws_cfs.append(["  Direct Taxes Paid", "", "", cfo.get("taxes_paid", 0.0)])
    apply_row_style(ws_cfs, ws_cfs.max_row, font=FONT_ITEM, num_format=NUM_FORMAT, align_cols={4: ALIGN_RIGHT})
    ws_cfs.append(["Net Cash Flow from Operating Activities (A)", "", "", cfo.get("net_cash_from_operating_activities", 0.0)])
    apply_row_style(ws_cfs, ws_cfs.max_row, fill=FILL_HIGHLIGHT, font=FONT_TOTAL, border=BORDER_THIN, num_format=NUM_FORMAT, align_cols={4: ALIGN_RIGHT})
    ws_cfs.append([])

    # B. Investing Activities
    ws_cfs.append(["B. CASH FLOW FROM INVESTING ACTIVITIES", "", "", ""])
    apply_row_style(ws_cfs, ws_cfs.max_row, fill=FILL_SUBTOTAL, font=FONT_SUB_HDR)
    for it in cfi.get("items", []):
        ws_cfs.append([f"  {it['label']}", "", "", it["amount"]])
        apply_row_style(ws_cfs, ws_cfs.max_row, font=FONT_ITEM, num_format=NUM_FORMAT, align_cols={4: ALIGN_RIGHT})
    ws_cfs.append(["Net Cash Flow from Investing Activities (B)", "", "", cfi.get("net_cash_from_investing_activities", 0.0)])
    apply_row_style(ws_cfs, ws_cfs.max_row, fill=FILL_HIGHLIGHT, font=FONT_TOTAL, border=BORDER_THIN, num_format=NUM_FORMAT, align_cols={4: ALIGN_RIGHT})
    ws_cfs.append([])

    # C. Financing Activities
    ws_cfs.append(["C. CASH FLOW FROM FINANCING ACTIVITIES", "", "", ""])
    apply_row_style(ws_cfs, ws_cfs.max_row, fill=FILL_SUBTOTAL, font=FONT_SUB_HDR)
    for it in cff.get("items", []):
        ws_cfs.append([f"  {it['label']}", "", "", it["amount"]])
        apply_row_style(ws_cfs, ws_cfs.max_row, font=FONT_ITEM, num_format=NUM_FORMAT, align_cols={4: ALIGN_RIGHT})
    ws_cfs.append(["Net Cash Flow from Financing Activities (C)", "", "", cff.get("net_cash_from_financing_activities", 0.0)])
    apply_row_style(ws_cfs, ws_cfs.max_row, fill=FILL_HIGHLIGHT, font=FONT_TOTAL, border=BORDER_THIN, num_format=NUM_FORMAT, align_cols={4: ALIGN_RIGHT})
    ws_cfs.append([])

    # Summary
    ws_cfs.append(["Net Increase / (Decrease) in Cash & Cash Equivalents (A + B + C)", "", "", cfs_sum.get("net_increase_decrease_cash", 0.0)])
    apply_row_style(ws_cfs, ws_cfs.max_row, font=FONT_SUB_HDR, border=BORDER_THIN, num_format=NUM_FORMAT, align_cols={4: ALIGN_RIGHT})
    ws_cfs.append(["Add: Cash & Cash Equivalents at Beginning of Year", "", "", cfs_sum.get("opening_cash_and_cash_equivalents", 0.0)])
    apply_row_style(ws_cfs, ws_cfs.max_row, font=FONT_ITEM, num_format=NUM_FORMAT, align_cols={4: ALIGN_RIGHT})
    ws_cfs.append(["Cash & Cash Equivalents at End of Year", "", "", cfs_sum.get("closing_cash_and_cash_equivalents", 0.0)])
    apply_row_style(ws_cfs, ws_cfs.max_row, fill=FILL_PRIMARY, font=Font(name="Segoe UI", size=10, bold=True, color="FFFFFF"), border=BORDER_TOTAL, num_format=NUM_FORMAT, align_cols={4: ALIGN_RIGHT})

    ws_cfs.column_dimensions["A"].width = 54
    ws_cfs.column_dimensions["B"].width = 5
    ws_cfs.column_dimensions["C"].width = 5
    ws_cfs.column_dimensions["D"].width = 24

    # =========================================================================
    # SHEET 5: NOTES TO ACCOUNTS
    # =========================================================================
    ws_notes = wb.create_sheet(title="Notes to Accounts")
    ws_notes.views.sheetView[0].showGridLines = True

    ws_notes.append([entity.get("name")])
    ws_notes.append(["NOTES FORMING PART OF THE FINANCIAL STATEMENTS"])
    ws_notes.append([])

    for note in notes_data:
        ws_notes.append([f"Note {note['note_number']}: {note['title']}", "", "", ""])
        apply_row_style(ws_notes, ws_notes.max_row, fill=FILL_PRIMARY, font=Font(name="Segoe UI", size=10, bold=True, color="FFFFFF"))
        
        ws_notes.append(["Ledger Account", "Account Code", "Debit", "Credit", "Net Amount"])
        apply_row_style(ws_notes, ws_notes.max_row, fill=FILL_SUBTOTAL, font=FONT_SUB_HDR, align_cols={1: ALIGN_LEFT, 2: ALIGN_CENTER, 3: ALIGN_RIGHT, 4: ALIGN_RIGHT, 5: ALIGN_RIGHT})

        for it in note.get("items", []):
            ws_notes.append([it["name"], it.get("code", ""), it.get("debit", 0.0), it.get("credit", 0.0), it.get("net", 0.0)])
            apply_row_style(ws_notes, ws_notes.max_row, font=FONT_ITEM, num_format=NUM_FORMAT, align_cols={2: ALIGN_CENTER, 3: ALIGN_RIGHT, 4: ALIGN_RIGHT, 5: ALIGN_RIGHT})

        ws_notes.append(["Total", "", "", "", note["total"]])
        apply_row_style(ws_notes, ws_notes.max_row, fill=FILL_HIGHLIGHT, font=FONT_TOTAL, border=BORDER_TOTAL, num_format=NUM_FORMAT, align_cols={5: ALIGN_RIGHT})
        ws_notes.append([])

    ws_notes.column_dimensions["A"].width = 46
    ws_notes.column_dimensions["B"].width = 16
    ws_notes.column_dimensions["C"].width = 18
    ws_notes.column_dimensions["D"].width = 18
    ws_notes.column_dimensions["E"].width = 20

    # =========================================================================
    # SHEET 6: PPE FIXED ASSETS SCHEDULE (AS 10)
    # =========================================================================
    ws_ppe = wb.create_sheet(title="PPE Asset Schedule")
    ws_ppe.views.sheetView[0].showGridLines = True

    ws_ppe.append([entity.get("name")])
    ws_ppe.append(["SCHEDULE OF PROPERTY, PLANT AND EQUIPMENT (AS 10)"])
    ws_ppe.append([])

    ws_ppe.append([
        "Asset Category", "Depr. Rate",
        "Gross Block (Opening)", "Additions", "Deletions", "Gross Block (Closing)",
        "Depreciation (Opening)", "For the Year", "Depreciation (Closing)",
        "Net Block (Closing)"
    ])
    apply_row_style(ws_ppe, ws_ppe.max_row, fill=FILL_PRIMARY, font=FONT_SECTION_HDR, align_cols={1: ALIGN_LEFT, 2: ALIGN_CENTER})

    for row in ppe.get("rows", []):
        gb = row["gross_block"]
        dep = row["depreciation"]
        nb = row["net_block"]
        ws_ppe.append([
            row["category"], row["depreciation_rate"],
            gb["opening"], gb["additions"], gb["deletions"], gb["closing"],
            dep["opening"], dep["for_the_year"], dep["closing"],
            nb["closing"]
        ])
        apply_row_style(ws_ppe, ws_ppe.max_row, font=FONT_ITEM, num_format=NUM_FORMAT, align_cols={2: ALIGN_CENTER})

    # PPE Totals
    t_gb = ppe.get("totals", {}).get("gross_block", {})
    t_dep = ppe.get("totals", {}).get("depreciation", {})
    t_nb = ppe.get("totals", {}).get("net_block", {})
    ws_ppe.append([
        "Total PPE", "-",
        t_gb.get("opening", 0.0), t_gb.get("additions", 0.0), t_gb.get("deletions", 0.0), t_gb.get("closing", 0.0),
        t_dep.get("opening", 0.0), t_dep.get("for_the_year", 0.0), t_dep.get("closing", 0.0),
        t_nb.get("closing", 0.0)
    ])
    apply_row_style(ws_ppe, ws_ppe.max_row, fill=FILL_HIGHLIGHT, font=FONT_TOTAL, border=BORDER_TOTAL, num_format=NUM_FORMAT, align_cols={2: ALIGN_CENTER})

    ws_ppe.column_dimensions["A"].width = 32
    for c in ["B", "C", "D", "E", "F", "G", "H", "I", "J"]:
        ws_ppe.column_dimensions[c].width = 18

    # =========================================================================
    # SHEET 7: 11 KEY FINANCIAL RATIOS
    # =========================================================================
    ws_rat = wb.create_sheet(title="Financial Ratios")
    ws_rat.views.sheetView[0].showGridLines = True

    ws_rat.append([entity.get("name")])
    ws_rat.append(["MANDATORY 11 KEY FINANCIAL RATIOS & BENCHMARK ANALYTICS"])
    ws_rat.append([])

    ws_rat.append(["Ratio Name", "Standard Formula", "Numerator (₹)", "Denominator (₹)", "Calculated Ratio", "Standard Benchmark", "Audit Status", "Analytical Interpretation"])
    apply_row_style(ws_rat, ws_rat.max_row, fill=FILL_PRIMARY, font=FONT_SECTION_HDR, align_cols={1: ALIGN_LEFT, 5: ALIGN_CENTER, 6: ALIGN_CENTER, 7: ALIGN_CENTER})

    for r in ratios:
        ws_rat.append([
            r["name"], r["formula"], r.get("numerator", 0.0), r.get("denominator", 0.0),
            f"{r['value']} {r['unit']}", r["benchmark"], r["status"], r["interpretation"]
        ])
        apply_row_style(ws_rat, ws_rat.max_row, font=FONT_ITEM, num_format=NUM_FORMAT, align_cols={5: ALIGN_CENTER, 6: ALIGN_CENTER, 7: ALIGN_CENTER})

    ws_rat.column_dimensions["A"].width = 34
    ws_rat.column_dimensions["B"].width = 42
    ws_rat.column_dimensions["C"].width = 18
    ws_rat.column_dimensions["D"].width = 18
    ws_rat.column_dimensions["E"].width = 18
    ws_rat.column_dimensions["F"].width = 18
    ws_rat.column_dimensions["G"].width = 14
    ws_rat.column_dimensions["H"].width = 46

    # =========================================================================
    # SHEET 8: AGING SCHEDULES
    # =========================================================================
    ws_age = wb.create_sheet(title="Aging Schedules")
    ws_age.views.sheetView[0].showGridLines = True

    ws_age.append([entity.get("name")])
    ws_age.append(["TRADE RECEIVABLES & TRADE PAYABLES AGING SCHEDULES"])
    ws_age.append([])

    # Receivables Aging
    ws_age.append(["1. Trade Receivables Aging Schedule", "", "", "", "", "", ""])
    apply_row_style(ws_age, ws_age.max_row, fill=FILL_PRIMARY, font=FONT_SECTION_HDR)

    ws_age.append(["Particulars", "< 6 Months", "6 Months - 1 Year", "1 - 2 Years", "2 - 3 Years", "> 3 Years", "Total"])
    apply_row_style(ws_age, ws_age.max_row, fill=FILL_SUBTOTAL, font=FONT_SUB_HDR, align_cols={1: ALIGN_LEFT})

    rec_aging = aging.get("receivables", {})
    for r in rec_aging.get("rows", []):
        ws_age.append([r["category"], r["less_6m"], r["6m_1y"], r["1y_2y"], r["2y_3y"], r["more_3y"], r["total"]])
        apply_row_style(ws_age, ws_age.max_row, font=FONT_ITEM, num_format=NUM_FORMAT)
    ws_age.append([])

    # Payables Aging
    ws_age.append(["2. Trade Payables Aging Schedule", "", "", "", "", ""])
    apply_row_style(ws_age, ws_age.max_row, fill=FILL_PRIMARY, font=FONT_SECTION_HDR)

    ws_age.append(["Particulars", "< 1 Year", "1 - 2 Years", "2 - 3 Years", "> 3 Years", "Total"])
    apply_row_style(ws_age, ws_age.max_row, fill=FILL_SUBTOTAL, font=FONT_SUB_HDR, align_cols={1: ALIGN_LEFT})

    pay_aging = aging.get("payables", {})
    for r in pay_aging.get("rows", []):
        ws_age.append([r["category"], r["less_1y"], r["1y_2y"], r["2y_3y"], r["more_3y"], r["total"]])
        apply_row_style(ws_age, ws_age.max_row, font=FONT_ITEM, num_format=NUM_FORMAT)

    ws_age.column_dimensions["A"].width = 46
    for c in ["B", "C", "D", "E", "F", "G"]:
        ws_age.column_dimensions[c].width = 18

    # =========================================================================
    # SHEET 9: SIGNIFICANT ACCOUNTING POLICIES
    # =========================================================================
    ws_pol = wb.create_sheet(title="Accounting Policies")
    ws_pol.views.sheetView[0].showGridLines = True

    ws_pol.append([entity.get("name")])
    ws_pol.append(["SIGNIFICANT ACCOUNTING POLICIES (FORMING PART OF ACCOUNTS)"])
    ws_pol.append([])

    ws_pol.append(["Accounting Standard", "Policy Subject", "Accounting Policy Description"])
    apply_row_style(ws_pol, ws_pol.max_row, fill=FILL_PRIMARY, font=FONT_SECTION_HDR)

    for pol in policies:
        ws_pol.append([pol["standard"], pol["title"], pol["policy"]])
        apply_row_style(ws_pol, ws_pol.max_row, font=FONT_ITEM)

    ws_pol.column_dimensions["A"].width = 36
    ws_pol.column_dimensions["B"].width = 30
    ws_pol.column_dimensions["C"].width = 75

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output
